import pandas as pd
import numpy as np
import logging
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def read_excel_data(filepath: str) -> pd.DataFrame:
    logging.info("[START] Reading Excel data")
    df = pd.read_excel(filepath, sheet_name='Sheet1', usecols="A:C")
    df.columns = df.columns.str.strip()
    logging.info("[END] Reading Excel data")
    return df

def calculate_daily_returns(df: pd.DataFrame) -> pd.DataFrame:
    logging.info("[START] Calculating daily returns")
    df['Nifty50_Return (%)'] = ((df['Nifty50'] / df['Nifty50'].shift(-1)) - 1) * 100
    df['Sensex_Return (%)'] = ((df['Sensex'] / df['Sensex'].shift(-1)) - 1) * 100
    logging.info("[END] Calculating daily returns")
    return df

def apply_border(cell):
    thin = Side(style='thin')
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def apply_number_format(cell):
    cell.number_format = '0.00'

def generate_weights() -> list:
    return [(150 - i * 10, -50 + i * 10) for i in range(21)]

def write_returns_to_excel(filepath: str, df: pd.DataFrame) -> None:
    logging.info("[START] Writing returns to Excel")
    wb = load_workbook(filepath)
    ws = wb['Sheet1']
    ws.cell(row=1, column=5, value="Nifty50 Return (%)").font = Font(bold=True)
    ws.cell(row=1, column=6, value="Sensex Return (%)").font = Font(bold=True)
    apply_border(ws.cell(row=1, column=5))
    apply_border(ws.cell(row=1, column=6))

    for idx, row in df.iterrows():
        row_index = idx + 2
        for col_offset, col_name in enumerate(['Nifty50_Return (%)', 'Sensex_Return (%)']):
            cell = ws.cell(row=row_index, column=5 + col_offset, value=row[col_name])
            apply_border(cell)
            apply_number_format(cell)

    wb.save(filepath)
    logging.info("[END] Writing returns to Excel")

def write_summary_statistics_to_excel(filepath: str, df_returns: pd.DataFrame) -> dict:
    logging.info("[START] Writing summary statistics to Excel")
    nifty_mean = df_returns['Nifty50_Return (%)'].mean()
    sensex_mean = df_returns['Sensex_Return (%)'].mean()
    nifty_std = df_returns['Nifty50_Return (%)'].std(ddof=0)
    sensex_std = df_returns['Sensex_Return (%)'].std(ddof=0)
    covariance = df_returns.cov().iloc[0, 1]
    risk_free_rate = 2 / 252

    wb = load_workbook(filepath)
    ws = wb['Sheet1']

    ws.cell(row=1, column=9, value="Nifty50").font = Font(bold=True)
    ws.cell(row=1, column=10, value="Sensex").font = Font(bold=True)
    apply_border(ws.cell(row=1, column=9))
    apply_border(ws.cell(row=1, column=10))

    labels = ['Mean', 'Standard Deviation', 'Covariance', 'Risk Free Rate']
    values_nifty = [nifty_mean, nifty_std, covariance, risk_free_rate]
    values_sensex = [sensex_mean, sensex_std, '', '']

    for i, label in enumerate(labels, start=2):
        ws.cell(row=i, column=8, value=label).font = Font(bold=True)
        apply_border(ws.cell(row=i, column=8))

        c1 = ws.cell(row=i, column=9, value=values_nifty[i - 2])
        c2 = ws.cell(row=i, column=10, value=values_sensex[i - 2] if values_sensex[i - 2] != '' else None)

        for cell in (c1, c2):
            if cell.value is not None:
                apply_border(cell)
                apply_number_format(cell)

    wb.save(filepath)
    logging.info("[END] Writing summary statistics to Excel")

    return {
        "nifty_mean": nifty_mean,
        "sensex_mean": sensex_mean,
        "nifty_std": nifty_std,
        "sensex_std": sensex_std,
        "covariance": covariance,
        "risk_free_rate": risk_free_rate
    }

def write_portfolio_weights(filepath: str) -> None:
    logging.info("[START] Writing portfolio weights")
    weights = generate_weights()
    wb = load_workbook(filepath)
    ws = wb['Sheet1']

    ws.cell(row=1, column=12, value="Nifty50 Weight (%)").font = Font(bold=True)
    ws.cell(row=1, column=13, value="Sensex Weight (%)").font = Font(bold=True)
    apply_border(ws.cell(row=1, column=12))
    apply_border(ws.cell(row=1, column=13))

    for i, (w1, w2) in enumerate(weights, start=2):
        c1 = ws.cell(row=i, column=12, value=w1)
        c2 = ws.cell(row=i, column=13, value=w2)
        for cell in (c1, c2):
            apply_border(cell)
            apply_number_format(cell)

    wb.save(filepath)
    logging.info("[END] Writing portfolio weights")

def write_portfolio_summary(filepath: str, stats: dict) -> None:
    logging.info("[START] Writing portfolio summary")
    wb = load_workbook(filepath)
    ws = wb['Sheet1']
    row_start = 2
    col_start = 15

    ws.cell(row=1, column=col_start, value="Portfolio Return").font = Font(bold=True)
    ws.cell(row=1, column=col_start + 1, value="Portfolio Risk").font = Font(bold=True)
    ws.cell(row=1, column=col_start + 2, value="Sharpe Ratio").font = Font(bold=True)

    for col in range(col_start, col_start + 3):
        apply_border(ws.cell(row=1, column=col))

    for i, (w1_raw, w2_raw) in enumerate(generate_weights()):
        w1 = w1_raw / 100
        w2 = w2_raw / 100

        port_return = (w1 * stats['nifty_mean']) + (w2 * stats['sensex_mean'])
        port_std = np.sqrt(
            (w1 ** 2) * (stats['nifty_std'] ** 2) +
            (w2 ** 2) * (stats['sensex_std'] ** 2) +
            2 * w1 * w2 * stats['covariance']
        )
        sharpe = (port_return - stats['risk_free_rate']) / port_std if port_std else 0

        for j, val in enumerate([port_return, port_std, sharpe]):
            cell = ws.cell(row=row_start + i, column=col_start + j, value=val)
            apply_border(cell)
            apply_number_format(cell)

    wb.save(filepath)
    logging.info("[END] Writing portfolio summary")

def write_cal(filepath: str, stats: dict) -> None:
    logging.info("[START] Writing Capital Allocation Line (CAL)")
    wb = load_workbook(filepath)
    ws = wb['Sheet1']
    col_start = 20

    ws.cell(row=1, column=col_start, value="Capital allocation line").font = Font(bold=True)
    ws.merge_cells(start_row=1, start_column=col_start, end_row=1, end_column=col_start + 2)
    ws.cell(row=2, column=col_start, value="Weight").font = Font(bold=True)
    ws.cell(row=2, column=col_start + 1, value="STD").font = Font(bold=True)
    ws.cell(row=2, column=col_start + 2, value="Return").font = Font(bold=True)

    for col in range(col_start, col_start + 3):
        apply_border(ws.cell(row=2, column=col))

    best_weight = max(generate_weights(),
                      key=lambda w: ((w[0]/100 * stats['nifty_mean']) + (w[1]/100 * stats['sensex_mean']) - stats['risk_free_rate']) /
                                    np.sqrt((w[0]/100)**2 * stats['nifty_std']**2 +
                                            (w[1]/100)**2 * stats['sensex_std']**2 +
                                            2 * (w[0]/100)*(w[1]/100)*stats['covariance']))
    w1, w2 = best_weight[0]/100, best_weight[1]/100
    best_ret = w1 * stats['nifty_mean'] + w2 * stats['sensex_mean']
    best_std = np.sqrt((w1 ** 2) * stats['nifty_std'] ** 2 +
                       (w2 ** 2) * stats['sensex_std'] ** 2 +
                       2 * w1 * w2 * stats['covariance'])

    for i, weight in enumerate([0, 1, 2], start=3):
        cal_ret = stats['risk_free_rate'] + weight * (best_ret - stats['risk_free_rate'])
        cal_std = weight * best_std
        ws.cell(row=i, column=col_start, value=f"{weight * 100}%")
        c1 = ws.cell(row=i, column=col_start + 1, value=cal_std)
        c2 = ws.cell(row=i, column=col_start + 2, value=cal_ret)

        for cell in (c1, c2):
            apply_border(cell)
            apply_number_format(cell)

    wb.save(filepath)
    logging.info("[END] Writing Capital Allocation Line (CAL)")

def main():
    logging.info("========== Portfolio Analysis Started ==========")
    path = "Efficient_Frontier.xlsx"
    df = read_excel_data(path)
    df = calculate_daily_returns(df)
    write_returns_to_excel(path, df)
    stats = write_summary_statistics_to_excel(path, df[['Nifty50_Return (%)', 'Sensex_Return (%)']].dropna())
    write_portfolio_weights(path)
    write_portfolio_summary(path, stats)
    write_cal(path, stats)
    logging.info("========== Portfolio Analysis Completed ==========")

if __name__ == "__main__":
    main()
